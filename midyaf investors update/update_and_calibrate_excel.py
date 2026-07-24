import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, coordinate_to_tuple
from openpyxl.cell.cell import MergedCell

def set_cell_safe(ws, coord, value):
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        row, col = coordinate_to_tuple(coord)
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left_cell.value = value
                return
    else:
        cell.value = value

def set_cell_rc_safe(ws, row, col, value):
    coord = get_column_letter(col) + str(row)
    set_cell_safe(ws, coord, value)

def update_sccc_sheet():
    print("Updating Collection Sheet SCCC.xlsx...")
    wb = openpyxl.load_workbook("Collection Sheet SCCC.xlsx")
    
    if "Intro" in wb.sheetnames:
        ws = wb["Intro"]
        set_cell_safe(ws, "B1", "Midyaf Royal & Event Hospitality Platform (مضياف)")
        set_cell_safe(ws, "B2", "Midyaf Engineering & Infrastructure Team")
        set_cell_safe(ws, "B3", "07/15/2026")
        
    if "Computing " in wb.sheetnames:
        ws = wb["Computing "]
        set_cell_safe(ws, "A3", 1)
        set_cell_safe(ws, "B3", "midyaf-core-api-prod")
        set_cell_safe(ws, "C3", "PRD - Production")
        set_cell_safe(ws, "D3", "Core API, Express/Node Server, Socket.IO Gateway")
        set_cell_safe(ws, "E3", 2)
        set_cell_safe(ws, "F3", 4)
        
        set_cell_safe(ws, "A4", 2)
        set_cell_safe(ws, "B4", "midyaf-ai-worker-prod")
        set_cell_safe(ws, "C4", "PRD - Production")
        set_cell_safe(ws, "D4", "Smart Guide (سيف ومنيره) AI Inference & WhatsApp Workers")
        set_cell_safe(ws, "E4", 2)
        set_cell_safe(ws, "F4", 8)
        
        set_cell_safe(ws, "A5", 3)
        set_cell_safe(ws, "B5", "midyaf-staging-dev")
        set_cell_safe(ws, "C5", "TST - Test/UAT")
        set_cell_safe(ws, "D5", "Staging, Pre-deployment QA & UAT Environment")
        set_cell_safe(ws, "E5", 1)
        set_cell_safe(ws, "F5", 2)
        
    if "Networking " in wb.sheetnames:
        ws = wb["Networking "]
        set_cell_safe(ws, "A4", 2)
        set_cell_safe(ws, "B4", "100 Mbps")
        set_cell_safe(ws, "C4", 1)
        set_cell_safe(ws, "E4", 500)
        set_cell_safe(ws, "D8", 2)
        
    if "Storage" in wb.sheetnames:
        ws = wb["Storage"]
        set_cell_safe(ws, "B3", 1)
        set_cell_safe(ws, "C3", 1000)
        
    if "RDS Database" in wb.sheetnames:
        ws = wb["RDS Database"]
        for r in range(1, ws.max_row+1):
            for c in range(1, ws.max_column+1):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    val = str(cell.value or '')
                    if 'Engine' in val or 'Database' in val:
                        set_cell_rc_safe(ws, r+1, 1, "PostgreSQL (Midyaf Core DB)")
                        set_cell_rc_safe(ws, r+1, 2, "PostgreSQL 16 High-Availability Dual-AZ")
                        set_cell_rc_safe(ws, r+1, 3, 1)
                        set_cell_rc_safe(ws, r+1, 4, "8 vCPU / 32 GB RAM")
                        set_cell_rc_safe(ws, r+1, 5, "500 GB SSD NVMe")
                        break
                    
    if "Security  " in wb.sheetnames:
        ws = wb["Security  "]
        set_cell_safe(ws, "B12", "Advanced Edition")
        set_cell_safe(ws, "B13", 5)
        set_cell_safe(ws, "B14", 20)
        set_cell_safe(ws, "B15", "500 GB")
        set_cell_safe(ws, "B16", "Yes (50 Piece)")
        set_cell_safe(ws, "B17", "200 GB")
        set_cell_safe(ws, "B18", "Yes (50 Piece)")
        set_cell_safe(ws, "B19", "Yes (10 Piece)")
        set_cell_safe(ws, "D21", "Enterprise Edition")
        
    wb.save("Collection Sheet SCCC.xlsx")
    print("Saved Collection Sheet SCCC.xlsx successfully.")

def update_topnet_sheet():
    print("Updating TopNet Clouding Questionaire.xlsx...")
    wb = openpyxl.load_workbook("TopNet Clouding Questionaire.xlsx")
    if "CLOUDING-REQUIREMENTS" in wb.sheetnames:
        ws = wb["CLOUDING-REQUIREMENTS"]
        vms = [
            (1, "Ubuntu", "16 GB", 8, "500 GB SSD", "Production Core API & Socket.IO App Server (Dual Instance)", "100 Mbps"),
            (2, "Ubuntu", "16 GB", 8, "500 GB SSD", "Smart AI Guide & WhatsApp API Background Worker Node", "100 Mbps"),
            (3, "Ubuntu", "32 GB", 16, "1024 GB NVMe", "PostgreSQL High-Availability Database Cluster (Master/Replica)", "200 Mbps"),
            (4, "Ubuntu", "8 GB", 4, "250 GB SSD", "Staging, Testing & Continuous Integration (CI/CD) Server", "50 Mbps"),
            (5, "Ubuntu", "8 GB", 4, "250 GB SSD", "Log Management, Prometheus Observability & Backup Vault", "50 Mbps"),
        ]
        for idx, os_val, vram, vcpu, storage, req, bw in vms:
            row = idx + 2
            set_cell_rc_safe(ws, row, 3, os_val)
            set_cell_rc_safe(ws, row, 4, vram)
            set_cell_rc_safe(ws, row, 5, vcpu)
            set_cell_rc_safe(ws, row, 6, storage)
            set_cell_rc_safe(ws, row, 7, req)
            set_cell_rc_safe(ws, row, 8, bw)
            
    wb.save("TopNet Clouding Questionaire.xlsx")
    print("Saved TopNet Clouding Questionaire.xlsx successfully.")

def update_financial_model():
    print("Updating and verifying Midyaf Financial Model (مضياف نموذج مالي.xlsx)...")
    wb = openpyxl.load_workbook("مضياف نموذج مالي.xlsx")
    
    if "Assumptions" in wb.sheetnames:
        ws = wb["Assumptions"]
        set_cell_safe(ws, "B129", "Cloud Hosting & Infrastructure (SCCC / TopNet Multi-AZ)")
        set_cell_safe(ws, "C129", 28800)
        set_cell_safe(ws, "D129", 43200)
        set_cell_safe(ws, "E129", 68400)
        set_cell_safe(ws, "F129", 68400)
        
        set_cell_safe(ws, "B109", "SMS / OTP Verification API (Taqnyat/Unifonic)")
        set_cell_safe(ws, "B114", "WhatsApp Business API & Bot Messaging (Meta/Twilio)")
        set_cell_safe(ws, "B115", "SCCC / TopNet Cloud Storage (OSS & Media Files per guest)")
        set_cell_safe(ws, "B117", "Cloud CDN & Edge Bandwidth (SCCC / TopNet Edge)")
        
    if "Income Statement" in wb.sheetnames:
        ws = wb["Income Statement"]
        set_cell_safe(ws, "A40", "Note on Cloud Infrastructure & Unit Economics: Cost of Revenue includes dynamic variable scaling per guest for WhatsApp API, SMS OTP, AI API usage (سيف ومنيره), and SCCC/TopNet Object Storage & CDN. Fixed Operating Expenses include dedicated SCCC / TopNet High-Availability Kubernetes and PostgreSQL RDS database clusters.")
        cell_40 = ws["A40"]
        if not isinstance(cell_40, MergedCell):
            cell_40.font = Font(name="Calibri", size=9, italic=True, color="555555")
        
    wb.save("مضياف نموذج مالي.xlsx")
    print("Saved مضياف نموذج مالي.xlsx successfully.")

if __name__ == "__main__":
    update_sccc_sheet()
    update_topnet_sheet()
    update_financial_model()
